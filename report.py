"""
report.py — Báo cáo hiệu quả KT hằng ngày (Excel + Google Drive).

Tối ưu:
- Dùng dữ liệu đã cào từ stats_data.py (2 account, sheets TI/Events/Spare/…/Solutions).
- Cửa sổ 1 ngày báo cáo: [0h hôm qua, 0h hôm nay) theo giờ VN.
- Mapping trạm → KT / region / địa bàn từ StationData.csv (GitHub).
- Sheet Cleaned Tickets thêm cột "Khu vực" = "new_district, new_province" ở cuối.
- Trạng thái đóng/mở suy ra từ Events (mới → cũ): OPEN / VOMS / CLOSED.
  CLOSED lấy Close Time = Create Time của "Pending for local team close" mới nhất;
  mỗi ticket chỉ tính 1 lần trong tháng. Ticket mở lại = đang OPEN nhưng lịch sử
  có VOMS hoặc local close. Status khác bỏ qua.
- Report đếm từ cùng tập Cleaned Tickets → khớp số ticket đóng tháng/ngày.
- Không ghi Excel lên disk (CCTS_Report); build in-memory rồi upload thẳng Drive.
"""
from __future__ import annotations

import os
import re
import sys
import warnings
from datetime import datetime, timedelta
from typing import Any
import json
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

from ccts_shared import VN_TZ, load_static_data_filtered, is_unmanaged_region

# Nạp file .env NẾU CÓ, để "python report.py" chạy standalone vẫn thấy được
# GOOGLE_CREDENTIALS_JSON, GOOGLE_TOKEN_JSON... Khi chạy qua app chính
# (app.py/main.py) mà module đó đã tự load_dotenv() rồi thì lệnh này chỉ là
# no-op (không ghi đè biến đã có sẵn trong môi trường thật, ví dụ trên Render).
try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Cấu hình Drive
# ---------------------------------------------------------------------------
DRIVE_FOLDER_ID = "1NSW79HL_ySob6xG08rZIKMkmm_UUjIzw"
CREDENTIALS_FILE = os.environ.get("GOOGLE_CREDENTIALS_JSON", "").strip()
TOKEN_FILE = os.environ.get("GOOGLE_TOKEN_JSON", "").strip()

_drive_service = None

# ---------------------------------------------------------------------------
# 1. Google Drive helpers
# ---------------------------------------------------------------------------

def get_drive_service():
    global _drive_service
    if _drive_service is not None:
        return _drive_service

    SCOPES = ["https://www.googleapis.com/auth/drive.file"]
    creds = None

    # Lấy giá trị token từ biến môi trường
    token_env = os.environ.get("GOOGLE_TOKEN_JSON", "").strip()

    # 1. Thử nạp Token đã lưu
    if token_env:
        if token_env.startswith("{"):
            # Nếu token_env là chuỗi JSON trực tiếp
            try:
                token_info = json.loads(token_env)
                creds = Credentials.from_authorized_user_info(token_info, SCOPES)
            except Exception as e:
                print(f"⚠️ Lỗi đọc GOOGLE_TOKEN_JSON từ chuỗi JSON: {e}")
        else:
            # Nếu token_env là đường dẫn file
            if os.path.exists(token_env):
                try:
                    creds = Credentials.from_authorized_user_file(token_env, SCOPES)
                except Exception as e:
                    print(f"⚠️ Lỗi đọc file token {token_env}: {e}")

    # 2. Nếu chưa có creds hoặc creds hết hạn
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(f"⚠️ Không thể refresh token, sẽ tiến hành đăng nhập lại: {e}")
                creds = None

        if not creds:
            # QUAN TRỌNG: trên server headless (Render, ...), KHÔNG có trình
            # duyệt và không ai bấm xác thực -> flow.run_local_server() sẽ
            # treo vô thời hạn chờ callback OAuth không bao giờ tới, khiến
            # toàn bộ tiến trình báo cáo bị "đứng im" mà không có exception
            # nào được raise (không lỗi hiển thị, chỉ đơn giản là không có
            # file nào được upload). Phải chặn nhánh này lại và fail nhanh,
            # rõ ràng thay vì treo.
            is_deployed = bool(
                os.environ.get("RENDER")
                or os.environ.get("IS_PULL_REQUEST") is not None
                or os.environ.get("DISABLE_OAUTH_INTERACTIVE_FLOW")
                or not os.environ.get("DISPLAY", None) and os.name != "nt" and not sys.stdin.isatty()
            )
            if is_deployed:
                print(
                    "❌ Refresh token thất bại và đang chạy trên môi trường không "
                    "tương tác (server/deploy) -> KHÔNG thể mở trình duyệt để "
                    "đăng nhập lại. Hãy tạo token.json mới ở máy local (chạy "
                    "get_drive_service() thủ công), rồi copy nội dung vào biến "
                    "môi trường GOOGLE_TOKEN_JSON trên server và redeploy."
                )
                return None

            raw_creds = os.environ.get("GOOGLE_CREDENTIALS_JSON", "").strip()
            if not raw_creds:
                print("⚠️ Chưa cấu hình biến môi trường GOOGLE_CREDENTIALS_JSON.")
                return None

            try:
                if raw_creds.startswith("{"):
                    creds_dict = json.loads(raw_creds)
                    flow = InstalledAppFlow.from_client_config(creds_dict, SCOPES)
                else:
                    if not os.path.exists(raw_creds):
                        print(f"⚠️ Không tìm thấy file credentials: {raw_creds}")
                        return None
                    flow = InstalledAppFlow.from_client_secrets_file(raw_creds, SCOPES)

                creds = flow.run_local_server(port=0)
            except Exception as e:
                print(f"❌ Lỗi khởi tạo OAuth flow: {e}")
                return None

        # 3. Lưu Token sau khi xác thực thành công
        save_path = "token.json"
        if token_env and not token_env.startswith("{"):
            save_path = token_env

        try:
            with open(save_path, "w", encoding="utf-8") as token_file:
                token_file.write(creds.to_json())
            print(f"   🔑 Đã lưu token đăng nhập vào file '{save_path}'")
        except Exception as e:
            print(f"⚠️ Không thể lưu token vào file: {e}")

    _drive_service = build("drive", "v3", credentials=creds)
    return _drive_service

def upload_bytes_to_drive(file_name: str, content: bytes, target_folder_id: str) -> None:
    """Upload Excel bytes thẳng lên Drive — không ghi file local."""
    try:
        service = get_drive_service()
        if not service:
            return

        query = (
            f"'{target_folder_id}' in parents and name = '{file_name}' "
            f"and trashed = false"
        )
        results = service.files().list(q=query, fields="files(id)").execute()
        items = results.get("files", [])

        file_metadata = {"name": file_name, "parents": [target_folder_id]}
        media = MediaIoBaseUpload(
            BytesIO(content),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True,
        )

        if items:
            file_id = items[0]["id"]
            updated = service.files().update(fileId=file_id, media_body=media).execute()
            print(f"   ☁️ Drive: Đã cập nhật đè '{file_name}' -> ID: {updated.get('id')}")
        else:
            new_file = service.files().create(
                body=file_metadata, media_body=media, fields="id"
            ).execute()
            print(f"   ☁️ Drive: Đã tải lên mới '{file_name}' -> ID: {new_file.get('id')}")
    except Exception as e:
        print(f"❌ Lỗi khi tải file lên Drive: {e}")


def get_or_create_date_folder(parent_folder_id: str, folder_name: str) -> str:
    try:
        service = get_drive_service()
        if not service:
            return parent_folder_id

        query = (
            f"'{parent_folder_id}' in parents and name = '{folder_name}' "
            f"and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
        )
        results = service.files().list(q=query, fields="files(id)").execute()
        folders = results.get("files", [])

        if folders:
            folder_id = folders[0]["id"]
            file_query = f"'{folder_id}' in parents and trashed = false"
            file_results = service.files().list(q=file_query, fields="files(id, name)").execute()
            for f in file_results.get("files", []):
                service.files().delete(fileId=f["id"]).execute()
            return folder_id

        folder_metadata = {
            "name": folder_name,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [parent_folder_id],
        }
        new_folder = service.files().create(body=folder_metadata, fields="id").execute()
        return new_folder.get("id")
    except Exception as e:
        print(f"❌ Lỗi Drive Folder: {e}")
        return parent_folder_id


# ---------------------------------------------------------------------------
# 2. Mapping StationData.csv (tech / region / địa bàn)
# ---------------------------------------------------------------------------

def _extract_core_station_code(code) -> str:
    try:
        from utils import extract_core_station_code
        core = extract_core_station_code(code)
        return core or ""
    except Exception:
        if code is None or (isinstance(code, float) and pd.isna(code)):
            return ""
        code_str = str(code).strip().upper()
        m = re.search(r"([A-Z]+\d+)", code_str)
        return m.group(1) if m else code_str


def load_station_maps_from_csv() -> tuple[dict, dict, dict]:
    """
    Trả (tech_map, region_map, admin_area_map).

    tech_map / region_map: load_static_data_filtered (HCM → "KV không quản lý").
    admin_area_map: {core_code: "new_district, new_province"} từ StationData.csv.
    """
    tech_map: dict = {}
    region_map: dict = {}
    admin_area_map: dict = {}

    try:
        _, tech_map, region_map, _, _ = load_static_data_filtered()
        tech_map = tech_map or {}
        region_map = region_map or {}
    except Exception as e:
        print(f"[report] Không load static filtered: {e}")

    try:
        from config import GITHUB_STATION_DATA_CSV_PATH
        from github_data_store import _fetch_github_csv

        df = _fetch_github_csv(GITHUB_STATION_DATA_CSV_PATH)
        col_map = {str(c).strip().lower(): c for c in df.columns}

        def _col(*names):
            for n in names:
                if n.lower() in col_map:
                    return col_map[n.lower()]
            return None

        c_code = _col("station_code")
        c_nd = _col("new_district")
        c_np = _col("new_province")
        c_tech = _col("technician", "tech", "kỹ thuật phụ trách")
        c_region = _col("region", "khu vực")

        if not c_code:
            print("[report] StationData.csv thiếu station_code — bỏ qua admin_area")
            return tech_map, region_map, admin_area_map

        for _, r in df.iterrows():
            raw_code = str(r.get(c_code) or "").strip()
            if not raw_code:
                continue
            core = _extract_core_station_code(raw_code)
            if not core:
                continue

            if c_tech and core not in tech_map:
                t = str(r.get(c_tech) or "").strip()
                if t:
                    tech_map[core] = t
            if c_region and core not in region_map:
                reg = str(r.get(c_region) or "").strip()
                if reg and not is_unmanaged_region(reg):
                    region_map[core] = reg

            nd = str(r.get(c_nd) or "").strip() if c_nd else ""
            np_ = str(r.get(c_np) or "").strip() if c_np else ""
            if nd or np_:
                admin_area_map[core] = ", ".join(x for x in (nd, np_) if x)

        print(
            f"[report] StationData: tech={len(tech_map)}, region={len(region_map)}, "
            f"admin_area={len(admin_area_map)}"
        )
    except Exception as e:
        print(f"[report] Lỗi đọc StationData.csv cho địa bàn: {e}")

    return tech_map, region_map, admin_area_map


def load_spare_name_map(spare_parts_file: str = "spareParts.xlsx") -> dict[str, str]:
    if not os.path.exists(spare_parts_file):
        print(f"[report] Không thấy {spare_parts_file} — dùng mã vật tư thô.")
        return {}
    try:
        df = pd.read_excel(spare_parts_file)
        if "Mã vật tư" in df.columns and "Tên vật tư" in df.columns:
            df["Mã vật tư"] = df["Mã vật tư"].astype(str).str.strip()
            return df.set_index("Mã vật tư")["Tên vật tư"].to_dict()
    except Exception as e:
        print(f"[report] Lỗi đọc {spare_parts_file}: {e}")
    return {}


# ---------------------------------------------------------------------------
# 3. Cửa sổ thời gian: [0h hôm qua, 0h hôm nay)
# ---------------------------------------------------------------------------

def report_time_window(now: datetime | None = None) -> tuple[datetime, datetime, str, str, int, int]:
    if now is None:
        now = datetime.now(VN_TZ)
    if now.tzinfo is not None:
        now = now.replace(tzinfo=None)

    end_today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    start_yesterday = end_today - timedelta(days=1)
    yesterday_str = start_yesterday.strftime("%d/%m")
    today_str = end_today.strftime("%d/%m")
    month = start_yesterday.month
    report_year = start_yesterday.year
    return start_yesterday, end_today, yesterday_str, today_str, month, report_year


# ---------------------------------------------------------------------------
# 4. Tổng hợp & định dạng Excel
# ---------------------------------------------------------------------------

def generate_report_dataframe(
    df_region_all: pd.DataFrame,
    df_region_closed: pd.DataFrame,
    start_yesterday: datetime,
    end_today: datetime,
    unclosed_statuses: list[str],
) -> pd.DataFrame:
    ktv_list = df_region_all["Kỹ thuật viên"].unique()
    summary_rows = []

    for ktv in ktv_list:
        if ktv == "Chưa phân công":
            continue

        df_ktv_all = df_region_all[df_region_all["Kỹ thuật viên"] == ktv].copy()
        df_ktv_closed_month = df_region_closed[df_region_closed["Kỹ thuật viên"] == ktv].copy()

        closed_tickets_yesterday = 0
        if not df_ktv_closed_month.empty and "Close Time_dt" in df_ktv_closed_month.columns:
            closed_tickets_yesterday = (
                (df_ktv_closed_month["Close Time_dt"] >= start_yesterday)
                & (df_ktv_closed_month["Close Time_dt"] < end_today)
            ).sum()

        if "Event_State" in df_ktv_all.columns:
            is_open_mask = df_ktv_all["Event_State"] == "OPEN"
        else:
            is_open_mask = df_ktv_all["Ticket Status"].isin(unclosed_statuses)
        df_ktv_open = df_ktv_all[
            is_open_mask & (df_ktv_all["Create Time_dt"] < end_today)
        ]

        is_backlog_bss = df_ktv_open["Charge Point ID"].astype(str).str.contains(
            "BSS", case=False, na=False
        )
        backlog_bss = int(is_backlog_bss.sum())
        backlog_trusac = int((~is_backlog_bss).sum())

        total_tickets_month = len(df_ktv_closed_month)
        if total_tickets_month == 0 and closed_tickets_yesterday == 0 and len(df_ktv_open) == 0:
            continue

        in_SLA = (
            (df_ktv_closed_month["SLA Status"] != "Overdue").sum()
            if "SLA Status" in df_ktv_closed_month.columns
            else 0
        )

        overdue_appt = overdue_spare = overdue_subjective_bss = overdue_subjective_trusac = 0

        if "Nguyên nhân Overdue" in df_ktv_closed_month.columns:
            overdue_appt = (
                df_ktv_closed_month["Nguyên nhân Overdue"]
                .astype(str)
                .str.contains("Appointment", case=False, na=False)
                .sum()
            )
            overdue_spare = (
                df_ktv_closed_month["Nguyên nhân Overdue"]
                .astype(str)
                .str.contains("Chờ vật tư", case=False, na=False)
                .sum()
            )
            list_reason = [
                "Lỗi chủ quan",
                "Ticket mở lại",
                "Lỗi chuyển trạng thái",
                "ASP trong hạn, đóng local team trễ",
            ]
            df_subjective = df_ktv_closed_month[
                df_ktv_closed_month["Nguyên nhân Overdue"].isin(list_reason)
            ]
            is_bss = df_subjective["Charge Point ID"].astype(str).str.contains(
                "BSS", case=False, na=False
            )
            overdue_subjective_bss = int(is_bss.sum())
            overdue_subjective_trusac = int((~is_bss).sum())

        total_remote = 0
        if "Handling type" in df_ktv_closed_month.columns:
            total_remote = len(
                df_ktv_closed_month[
                    df_ktv_closed_month["Handling type"]
                    .astype(str)
                    .str.contains("Từ xa", case=False, na=False)
                ]
            )
        total_onsite = total_tickets_month - total_remote

        efficiency = (
            (in_SLA + overdue_appt + overdue_spare) / total_tickets_month
            if total_tickets_month > 0
            else 0
        )
        overdue_pct = (
            (overdue_subjective_trusac + overdue_subjective_bss) / total_tickets_month
            if total_tickets_month > 0
            else 0
        )

        summary_rows.append(
            {
                "Tên kỹ thuật": ktv,
                "Số ticket đã đóng": closed_tickets_yesterday,
                "Tồn - Trụ sạc": backlog_trusac,
                "Tồn - BSS": backlog_bss,
                "Tổng ticket - Từ xa": total_remote,
                "Tổng ticket - Tại trạm": total_onsite,
                "Xử lý trong hạn": in_SLA,
                "Do hẹn khách": overdue_appt,
                "Do chờ vật tư": overdue_spare,
                "Do chủ quan - Trụ sạc": overdue_subjective_trusac,
                "Do chủ quan - BSS": overdue_subjective_bss,
                "Hiệu suất": efficiency,
                "Phần trăm Overdue": overdue_pct,
            }
        )

    return pd.DataFrame(summary_rows)


def check_missing_spare_parts(
    df_region_closed: pd.DataFrame,
    df_events: pd.DataFrame,
    df_spare: pd.DataFrame,
    start_yesterday: datetime,
    end_today: datetime,
) -> pd.DataFrame:
    if df_events.empty or "Ticket Status" not in df_events.columns:
        return pd.DataFrame()
    spare_event_tickets = set(
        df_events[df_events["Ticket Status"] == "Pending for spare parts"]["Ticket ID"].unique()
    )
    spare_recorded_tickets = (
        set(df_spare["Ticket ID"].unique()) if not df_spare.empty else set()
    )

    if df_region_closed.empty or "Close Time_dt" not in df_region_closed.columns:
        return pd.DataFrame()
    df_target = df_region_closed[
        (df_region_closed["Close Time_dt"] >= start_yesterday)
        & (df_region_closed["Close Time_dt"] < end_today)
    ].copy()
    if df_target.empty:
        return pd.DataFrame()

    is_missing = df_target["Ticket ID"].apply(
        lambda x: x in spare_event_tickets and x not in spare_recorded_tickets
    )
    df_warning = df_target[is_missing]

    output_cols = [
        "Kỹ thuật viên",
        "Ticket ID",
        "Station Code",
        "Charge Point ID",
        "Problem Description",
    ]
    existing = [c for c in output_cols if c in df_warning.columns]
    res = df_warning[existing].copy()
    if "Kỹ thuật viên" in res.columns:
        res = res.rename(columns={"Kỹ thuật viên": "Tên kỹ thuật"})
    return res


def get_replaced_parts_yesterday(
    df_month_all: pd.DataFrame,
    df_spare: pd.DataFrame,
    start_yesterday: datetime,
    end_today: datetime,
) -> pd.DataFrame:
    if df_spare.empty:
        return pd.DataFrame()
    df_spare = df_spare.copy()
    df_spare["Create Time_dt"] = pd.to_datetime(
        df_spare["Create Time"], format="%Y-%m-%d %H:%M:%S", errors="coerce"
    )
    df_spare_yesterday = df_spare[
        (df_spare["Create Time_dt"] >= start_yesterday)
        & (df_spare["Create Time_dt"] < end_today)
    ]
    if df_spare_yesterday.empty:
        return pd.DataFrame()

    groupby_cols = ["Ticket ID", "Material Name (English)", "Material Code"]
    if "Processor" in df_spare_yesterday.columns:
        groupby_cols.append("Processor")

    grouped = (
        df_spare_yesterday.groupby(groupby_cols).size().reset_index(name="Usage Quantity")
    )
    ktv_map = df_month_all.set_index("Ticket ID")["Kỹ thuật viên"].to_dict()
    station_map = df_month_all.set_index("Ticket ID")["Station Code"].to_dict()

    grouped["Tên kỹ thuật"] = grouped["Ticket ID"].map(ktv_map)
    grouped["Station Code"] = grouped["Ticket ID"].map(station_map)
    grouped = grouped[
        pd.notna(grouped["Tên kỹ thuật"]) & (grouped["Tên kỹ thuật"] != "Chưa phân công")
    ]
    grouped["Usage Quantity"] = 1

    if "Processor" in grouped.columns:
        grouped = grouped.rename(columns={"Processor": "Tài khoản"})
    else:
        grouped["Tài khoản"] = ""

    return grouped[
        [
            "Tên kỹ thuật",
            "Ticket ID",
            "Station Code",
            "Material Name (English)",
            "Material Code",
            "Usage Quantity",
            "Tài khoản",
        ]
    ]


def export_multisheet_excel(
    df_region_data: pd.DataFrame,
    report_df: pd.DataFrame | None,
    df_warning: pd.DataFrame | None,
    df_parts: pd.DataFrame | None,
    yesterday_str: str,
    today_str: str,
    month: int,
    is_unassigned: bool = False,
) -> bytes:
    wb = Workbook()
    font_title = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_blue_dark = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_blue_light = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_orange_header = PatternFill(start_color="D66011", end_color="D66011", fill_type="solid")
    fill_green_header = PatternFill(start_color="1E6B34", end_color="1E6B34", fill_type="solid")
    fill_total_row = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_red_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    font_body = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    if not is_unassigned:
        ws_rep = wb.active
        ws_rep.title = "Report"
        ws_rep.views.sheetView[0].showGridLines = True

        ws_rep["A1"] = "Tên kỹ thuật"
        ws_rep["B1"] = f"HIỆU QUẢ TRONG NGÀY {yesterday_str}"
        ws_rep["B2"] = "Số ticket đã đóng"
        ws_rep["C2"] = f"Số ticket còn tồn {today_str}"
        ws_rep["C3"] = "Trụ sạc"
        ws_rep["D3"] = "BSS"

        ws_rep["E1"] = f"HIỆU QUẢ XỬ LÝ (THÁNG {month})"
        ws_rep["E2"] = "Tổng ticket"
        ws_rep["E3"] = "Từ xa"
        ws_rep["F3"] = "Tại trạm"
        ws_rep["G2"] = "Xử lý trong hạn"

        ws_rep["H1"] = "Overdue"
        ws_rep["H2"] = "Do hẹn khách"
        ws_rep["I2"] = "Do chờ vật tư"
        ws_rep["J2"] = "Do chủ quan"
        ws_rep["J3"] = "Trụ sạc"
        ws_rep["K3"] = "BSS"

        ws_rep["L1"] = "Hiệu suất"
        ws_rep["M1"] = "Phần trăm Overdue"

        ws_rep.merge_cells("A1:A3")
        ws_rep.merge_cells("B1:D1")
        ws_rep.merge_cells("B2:B3")
        ws_rep.merge_cells("C2:D2")
        ws_rep.merge_cells("E1:G1")
        ws_rep.merge_cells("E2:F2")
        ws_rep.merge_cells("G2:G3")
        ws_rep.merge_cells("H1:K1")
        ws_rep.merge_cells("H2:H3")
        ws_rep.merge_cells("I2:I3")
        ws_rep.merge_cells("J2:K2")
        ws_rep.merge_cells("L1:L3")
        ws_rep.merge_cells("M1:M3")

        for r in [1, 2, 3]:
            for c in range(1, 14):
                cell = ws_rep.cell(row=r, column=c)
                cell.font = font_title
                cell.alignment = align_center
                cell.border = border_thin
                if c in [2, 3, 4]:
                    cell.fill = fill_orange_header
                elif c in [8, 9, 10, 11]:
                    cell.fill = fill_red_header
                else:
                    cell.fill = fill_blue_dark if r == 1 else fill_blue_light

        if report_df is not None and not report_df.empty:
            for row_data in report_df.itertuples(index=False):
                ws_rep.append(list(row_data))

            sum_closed_yesterday = report_df["Số ticket đã đóng"].sum()
            sum_backlog_trusac = report_df["Tồn - Trụ sạc"].sum()
            sum_backlog_bss = report_df["Tồn - BSS"].sum()
            sum_remote = report_df["Tổng ticket - Từ xa"].sum()
            sum_onsite = report_df["Tổng ticket - Tại trạm"].sum()
            sum_in_sla = report_df["Xử lý trong hạn"].sum()
            sum_appt = report_df["Do hẹn khách"].sum()
            sum_spare = report_df["Do chờ vật tư"].sum()
            sum_sub_trusac = report_df["Do chủ quan - Trụ sạc"].sum()
            sum_sub_bss = report_df["Do chủ quan - BSS"].sum()

            total_tickets_month = sum_remote + sum_onsite
            total_efficiency = (
                (sum_in_sla + sum_appt + sum_spare) / total_tickets_month
                if total_tickets_month > 0
                else 0
            )
            total_overdue_pct = (
                (sum_sub_trusac + sum_sub_bss) / total_tickets_month
                if total_tickets_month > 0
                else 0
            )

            ws_rep.append(
                [
                    "Tổng",
                    sum_closed_yesterday,
                    sum_backlog_trusac,
                    sum_backlog_bss,
                    sum_remote,
                    sum_onsite,
                    sum_in_sla,
                    sum_appt,
                    sum_spare,
                    sum_sub_trusac,
                    sum_sub_bss,
                    total_efficiency,
                    total_overdue_pct,
                ]
            )

        max_report_row = ws_rep.max_row

        for r in range(4, max_report_row):
            ws_rep.row_dimensions[r].height = 22
            for c in range(1, 14):
                cell = ws_rep.cell(row=r, column=c)
                cell.border = border_thin
                if c == 1:
                    cell.font = font_body
                    cell.alignment = align_left
                elif c in [12, 13]:
                    cell.font = font_bold
                    cell.alignment = align_center
                    cell.number_format = "0.0%"
                else:
                    cell.font = font_body
                    cell.alignment = align_center

        if report_df is not None and not report_df.empty:
            ws_rep.row_dimensions[max_report_row].height = 24
            for c in range(1, 14):
                cell = ws_rep.cell(row=max_report_row, column=c)
                cell.font = font_bold
                cell.border = border_thin
                cell.fill = fill_total_row
                if c == 1:
                    cell.alignment = align_left
                elif c in [12, 13]:
                    cell.alignment = align_center
                    cell.number_format = "0.0%"
                else:
                    cell.alignment = align_center

        if max_report_row >= 5:
            color_scale_eff = ColorScaleRule(
                start_type="min",
                start_color="FF4D4D",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFC000",
                end_type="max",
                end_color="00B050",
            )
            ws_rep.conditional_formatting.add(f"L4:L{max_report_row}", color_scale_eff)
            color_scale_ovd = ColorScaleRule(
                start_type="num",
                start_value=0.0,
                start_color="FFFFFF",
                mid_type="num",
                mid_value=0.2,
                mid_color="FF4D4D",
                end_type="num",
                end_value=1.0,
                end_color="7030A0",
            )
            ws_rep.conditional_formatting.add(f"M4:M{max_report_row}", color_scale_ovd)

        current_cursor_row = max_report_row

        if df_warning is not None and not df_warning.empty:
            start_warn_row = current_cursor_row + 3
            ws_rep.cell(
                row=start_warn_row,
                column=1,
                value=f"Các ticket chưa add vật tư {yesterday_str}",
            ).font = Font(name="Segoe UI", size=12, bold=True, color="C00000")
            for idx, h_text in enumerate(list(df_warning.columns), start=1):
                cell = ws_rep.cell(row=start_warn_row + 1, column=idx, value=h_text)
                cell.font = font_title
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.alignment = align_center
                cell.border = border_thin
            for r_idx, row_data in enumerate(
                df_warning.itertuples(index=False), start=start_warn_row + 2
            ):
                ws_rep.row_dimensions[r_idx].height = 22
                for c_idx, val in enumerate(list(row_data), start=1):
                    cell = ws_rep.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = font_body
                    cell.border = border_thin
                    cell.alignment = align_left if c_idx in [1, 5] else align_center
                    if c_idx == 2:
                        cell.number_format = "@"
            current_cursor_row = ws_rep.max_row

        if df_parts is not None and not df_parts.empty:
            start_parts_row = current_cursor_row + 3
            ws_rep.cell(
                row=start_parts_row,
                column=1,
                value=f"Các kỹ thuật đã thay vật tư trong ngày {yesterday_str}",
            ).font = Font(name="Segoe UI", size=12, bold=True, color="1E6B34")
            for idx, h_text in enumerate(list(df_parts.columns), start=1):
                cell = ws_rep.cell(row=start_parts_row + 1, column=idx, value=h_text)
                cell.font = font_title
                cell.fill = fill_green_header
                cell.alignment = align_center
                cell.border = border_thin
            for r_idx, row_data in enumerate(
                df_parts.itertuples(index=False), start=start_parts_row + 2
            ):
                ws_rep.row_dimensions[r_idx].height = 22
                for c_idx, val in enumerate(list(row_data), start=1):
                    cell = ws_rep.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = font_body
                    cell.border = border_thin
                    cell.alignment = align_left if c_idx in [1, 4] else align_center
                    if c_idx == 2:
                        cell.number_format = "@"

        ws_rep.column_dimensions["A"].width = 25
        for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
            ws_rep.column_dimensions[col_letter].width = 16
        ws_rep.row_dimensions[1].height = 26
        ws_rep.row_dimensions[2].height = 22
        ws_rep.row_dimensions[3].height = 22

        ws_dtl = wb.create_sheet(title="Cleaned Tickets")
    else:
        ws_dtl = wb.active
        ws_dtl.title = "Unassigned Tickets"

    ws_dtl.views.sheetView[0].showGridLines = True
    ws_dtl.append(list(df_region_data.columns))
    for row in df_region_data.itertuples(index=False, name=None):
        ws_dtl.append(row)

    for col_idx in range(1, len(df_region_data.columns) + 1):
        cell = ws_dtl.cell(row=1, column=col_idx)
        cell.font = font_title
        cell.fill = fill_blue_dark
        cell.alignment = align_center
        cell.border = border_thin

    col_names = list(df_region_data.columns)
    ticket_id_idx = col_names.index("Ticket ID") + 1 if "Ticket ID" in col_names else -1
    sla_status_idx = col_names.index("SLA Status") + 1 if "SLA Status" in col_names else -1
    nguyen_nhan_idx = (
        col_names.index("Nguyên nhân Overdue") + 1 if "Nguyên nhân Overdue" in col_names else -1
    )

    for row_idx in range(2, ws_dtl.max_row + 1):
        is_row_overdue = False
        is_yellow = False

        if sla_status_idx != -1:
            is_row_overdue = ws_dtl.cell(row=row_idx, column=sla_status_idx).value == "Overdue"

        if is_row_overdue and nguyen_nhan_idx != -1:
            reason_val = ws_dtl.cell(row=row_idx, column=nguyen_nhan_idx).value or ""
            if str(reason_val).strip() != "Lỗi chủ quan":
                is_yellow = True

        for col_idx in range(1, len(df_region_data.columns) + 1):
            cell = ws_dtl.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = border_thin
            cell.alignment = (
                align_left
                if col_names[col_idx - 1]
                in [
                    "Problem Description",
                    "Address",
                    "Solution",
                    "Giải trình",
                    "Nguyên nhân Overdue",
                    "Khu vực",
                ]
                else align_center
            )
            if col_idx == ticket_id_idx:
                cell.number_format = "@"

            if is_row_overdue:
                if is_yellow:
                    cell.fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )
                    if col_idx == sla_status_idx:
                        cell.font = Font(name="Segoe UI", size=10, color="B8860B", bold=True)
                else:
                    cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                    if col_idx == sla_status_idx:
                        cell.font = Font(name="Segoe UI", size=10, color="9C0006", bold=True)

    for col in ws_dtl.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col)
        adjusted_width = min(max(max_len + 3, 12), 42)
        if col[0].value in [
            "Problem Description",
            "Address",
            "Solution",
            "Giải trình",
            "Nguyên nhân Overdue",
            "Khu vực",
        ]:
            adjusted_width = 38
        ws_dtl.column_dimensions[col_letter].width = adjusted_width
    ws_dtl.row_dimensions[1].height = 28
    for r in range(2, ws_dtl.max_row + 1):
        ws_dtl.row_dimensions[r].height = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 5. Xử lý chính — nhận dữ liệu đã cào từ stats_data
# ---------------------------------------------------------------------------

def _clean_tid(val) -> str:
    if pd.isna(val):
        return ""
    s = str(val)
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip()


def process_and_clean_data(
    dfs_es: dict[str, pd.DataFrame] | None,
    dfs_its: dict[str, pd.DataFrame] | None,
    *,
    spare_parts_file: str = "spareParts.xlsx",
    excluded_prefixes: list[str] | None = None,
    now: datetime | None = None,
) -> pd.DataFrame | None:
    """
    Nhận dict sheet từ 2 account (đã cào bởi stats_data / CCTSClient),
    chuẩn hoá, xuất Excel theo vùng + upload Drive.

    Cửa sổ ngày: [0h hôm qua, 0h hôm nay).
    Mapping: StationData.csv (tech, region, new_district+new_province).
    """
    start_yesterday, end_today, yesterday_str, today_str, month, report_year = report_time_window(
        now
    )
    print(
        f"--- BÁO CÁO NGÀY {yesterday_str} "
        f"([{start_yesterday} → {end_today})) — THÁNG {month}/{report_year} ---"
    )

    if not dfs_es and not dfs_its:
        print("❌ Không có dữ liệu đầu vào từ cả 2 tài khoản.")
        return None

    required_sheets = [
        "Ticket Information",
        "Appointment",
        "Events Record",
        "Solutions",
        "Spare Parts Record",
        "Additional information",
    ]
    if not dfs_es:
        dfs_es = {s: pd.DataFrame() for s in required_sheets}
    if not dfs_its:
        dfs_its = {s: pd.DataFrame() for s in required_sheets}
    for s in required_sheets:
        dfs_es.setdefault(s, pd.DataFrame())
        dfs_its.setdefault(s, pd.DataFrame())

    print("1. Hợp nhất DataFrame từ 2 tài khoản...")
    df_raw = pd.concat(
        [
            dfs_es.get("Ticket Information", pd.DataFrame()),
            dfs_its.get("Ticket Information", pd.DataFrame()),
        ],
        ignore_index=True,
    )
    df_appt = pd.concat(
        [dfs_es.get("Appointment", pd.DataFrame()), dfs_its.get("Appointment", pd.DataFrame())],
        ignore_index=True,
    )
    df_events = pd.concat(
        [
            dfs_es.get("Events Record", pd.DataFrame()),
            dfs_its.get("Events Record", pd.DataFrame()),
        ],
        ignore_index=True,
    )
    df_sol = pd.concat(
        [dfs_es.get("Solutions", pd.DataFrame()), dfs_its.get("Solutions", pd.DataFrame())],
        ignore_index=True,
    )
    df_spare = pd.concat(
        [
            dfs_es.get("Spare Parts Record", pd.DataFrame()),
            dfs_its.get("Spare Parts Record", pd.DataFrame()),
        ],
        ignore_index=True,
    )
    df_info = pd.concat(
        [
            dfs_es.get("Additional information", pd.DataFrame()),
            dfs_its.get("Additional information", pd.DataFrame()),
        ],
        ignore_index=True,
    )

    if df_raw.empty:
        print("❌ Sheet 'Ticket Information' rỗng sau khi gộp.")
        return None

    for df in (df_raw, df_appt, df_events, df_sol, df_spare, df_info):
        if not df.empty and "Ticket ID" in df.columns:
            df["Ticket ID"] = df["Ticket ID"].apply(_clean_tid)

    dict_spare_names = load_spare_name_map(spare_parts_file)
    unclosed_statuses = [
        "Open",
        "Appointment",
        "Pending for spare parts",
        "Pending for ASP close",
    ]
    voms_status = "Pending for VOMS confirm"
    local_close_status = "Pending for local team close"

    # Event-derived state (Events newest → oldest).
    # OPEN: gặp unclosed → đang mở; quét tiếp, gặp VOMS/local close → mở lại.
    # VOMS: không đóng, không mở.
    # CLOSED: local close mới nhất = Close Time; mỗi ticket 1 lần/tháng.
    # Status khác: bỏ qua.
    close_time_map: dict = {}
    first_close_map: dict = {}
    first_asp_map: dict = {}
    last_asp_map: dict = {}
    reopened_tickets_set: set = set()
    event_state_map: dict = {}

    if not df_events.empty and "Ticket Status" in df_events.columns:
        df_ev = df_events.copy()
        df_ev["Ticket ID"] = df_ev["Ticket ID"].apply(_clean_tid)
        df_ev["Event Time_dt"] = pd.to_datetime(
            df_ev["Create Time"], format="%Y-%m-%d %H:%M:%S", errors="coerce"
        )
        df_ev = df_ev.dropna(subset=["Event Time_dt"])
        df_ev = df_ev.sort_values(
            ["Ticket ID", "Event Time_dt"], ascending=[True, False]
        )

        for tid, group in df_ev.groupby("Ticket ID", sort=False):
            state = None
            closes: list = []
            for _, ev in group.iterrows():
                st = str(ev["Ticket Status"]).strip()
                t = ev["Event Time_dt"]

                if state is None:
                    if st in unclosed_statuses:
                        state = "OPEN"
                        continue
                    if st == voms_status:
                        state = "VOMS"
                        break
                    if st == local_close_status:
                        state = "CLOSED"
                        closes.append(t)
                        continue
                    continue

                if state == "OPEN":
                    if st == voms_status or st == local_close_status:
                        reopened_tickets_set.add(tid)
                        break
                    continue

                if state == "CLOSED" and st == local_close_status:
                    closes.append(t)

            if state is None:
                state = "UNKNOWN"
            event_state_map[tid] = state
            if closes:
                close_time_map[tid] = closes[0]
                first_close_map[tid] = closes[-1]
                if len(closes) > 1:
                    reopened_tickets_set.add(tid)

        df_asp_events = df_ev[df_ev["Ticket Status"] == "Pending for ASP close"].copy()
        if not df_asp_events.empty:
            df_asp_events = df_asp_events.sort_values(
                ["Ticket ID", "Event Time_dt"], ascending=[True, True]
            )
            asp_times_dict = (
                df_asp_events.groupby("Ticket ID")["Event Time_dt"].apply(list).to_dict()
            )
            first_asp_map = {
                tid: times[0] for tid, times in asp_times_dict.items() if times
            }
            last_asp_map = {
                tid: times[-1]
                for tid, times in asp_times_dict.items()
                if len(times) > 1
            }

    df_raw["ASP close_dt"] = df_raw["Ticket ID"].map(first_asp_map)
    df_raw["Last_ASP_dt"] = df_raw["Ticket ID"].map(last_asp_map)
    df_raw["Close Time_dt"] = df_raw["Ticket ID"].map(close_time_map)
    df_raw["First Local Close_dt"] = df_raw["Ticket ID"].map(first_close_map)
    df_raw["Event_State"] = df_raw["Ticket ID"].map(event_state_map).fillna("UNKNOWN")
    df_raw["ASP close"] = (
        df_raw["ASP close_dt"].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
    )
    df_raw["Create Time_dt"] = pd.to_datetime(
        df_raw["Create Time"], format="%Y-%m-%d %H:%M:%S", errors="coerce"
    )

    _unknown = df_raw["Event_State"] == "UNKNOWN"
    if _unknown.any():
        _ti_open = df_raw["Ticket Status"].isin(unclosed_statuses)
        _ti_voms = df_raw["Ticket Status"].astype(str).str.strip() == voms_status
        df_raw.loc[_unknown & _ti_open, "Event_State"] = "OPEN"
        df_raw.loc[_unknown & _ti_voms, "Event_State"] = "VOMS"

    is_closed_in_month = (
        (df_raw["Event_State"] == "CLOSED")
        & df_raw["Close Time_dt"].notna()
        & (df_raw["Close Time_dt"].dt.month == month)
        & (df_raw["Close Time_dt"].dt.year == report_year)
    )
    is_open_backlog = df_raw["Event_State"] == "OPEN"

    df_month_all = df_raw[is_closed_in_month | is_open_backlog].copy()
    if df_month_all.empty:
        print(f"⚠️ Không có ticket hợp lệ cho tháng {month}/{report_year}.")
        return pd.DataFrame()

    if "Problem Description" in df_month_all.columns:
        df_month_all = df_month_all[
            ~df_month_all["Problem Description"].astype(str).str.startswith("BSS.No2")
        ]

    if excluded_prefixes and "Station Code" in df_month_all.columns:
        df_month_all = df_month_all[
            df_month_all["Station Code"].apply(
                lambda x: not any(p in str(x) for p in excluded_prefixes)
            )
        ]

    appt_dict = (
        df_appt.drop_duplicates(subset=["Ticket ID"], keep="last")
        .set_index("Ticket ID")["Detail"]
        .to_dict()
        if not df_appt.empty and "Detail" in df_appt.columns
        else {}
    )
    spare_parts_tickets = (
        set(
            df_events[df_events["Ticket Status"] == "Pending for spare parts"][
                "Ticket ID"
            ].unique()
        )
        if not df_events.empty and "Ticket Status" in df_events.columns
        else set()
    )

    ticket_to_part_str: dict[str, str] = {}
    if not df_spare.empty and "Ticket ID" in df_spare.columns and "Material Code" in df_spare.columns:
        if "Material type" in df_spare.columns:
            df_spare_good = df_spare[
                df_spare["Material type"].astype(str).str.strip().str.lower() == "good parts"
            ]
        else:
            df_spare_good = df_spare
        for _, row in df_spare_good.iterrows():
            tid = str(row["Ticket ID"]).strip()
            mat_code = str(row["Material Code"]).strip()
            mat_name = dict_spare_names.get(mat_code)
            part_display = (
                str(mat_name).strip()
                if pd.notna(mat_name) and str(mat_name).strip() not in ("", "nan")
                else mat_code
            )
            if tid in ticket_to_part_str:
                if part_display not in ticket_to_part_str[tid]:
                    ticket_to_part_str[tid] += f", {part_display}"
            else:
                ticket_to_part_str[tid] = part_display

    info_dict = (
        df_info.drop_duplicates(subset=["Ticket ID"], keep="last")
        .set_index("Ticket ID")["Handling type"]
        .to_dict()
        if not df_info.empty and "Handling type" in df_info.columns
        else {}
    )

    tech_map, region_map, admin_area_map = load_station_maps_from_csv()

    df_month_all["Clean_Station_Code"] = df_month_all["Station Code"].apply(
        _extract_core_station_code
    )
    df_month_all["Kỹ thuật viên"] = (
        df_month_all["Clean_Station_Code"].map(tech_map).fillna("Chưa phân công")
    )
    df_month_all["Region_op"] = (
        df_month_all["Clean_Station_Code"].map(region_map).fillna("Chưa phân công")
    )
    df_month_all["Khu vực"] = (
        df_month_all["Clean_Station_Code"].map(admin_area_map).fillna("")
    )

    perm_sol_dict: dict = {}
    custom_sol_dict: dict = {}
    if not df_sol.empty and "Ticket ID" in df_sol.columns and "Solutions Type" in df_sol.columns:
        for _, row in df_sol.iterrows():
            tid = str(row["Ticket ID"]).strip()
            sol_type = str(row["Solutions Type"]).strip()
            desc = str(row.get("Solution Description", "") or "")
            if pd.isna(row.get("Solution Description")):
                desc = ""
            if sol_type == "Permanent solution":
                perm_sol_dict[tid] = desc
            elif sol_type == "Customer solutions":
                custom_sol_dict[tid] = desc

    df_month_all["Solution"] = df_month_all["Ticket ID"].map(
        lambda x: perm_sol_dict.get(str(x).strip(), "")
    )
    df_month_all["Giải trình"] = df_month_all["Ticket ID"].map(
        lambda x: custom_sol_dict.get(str(x).strip(), "")
    )
    df_month_all["Handling type"] = df_month_all["Ticket ID"].map(info_dict).fillna("")
    if "Create Time" in df_month_all.columns:
        df_month_all["Create Time"] = df_month_all["Create Time"].astype(str)

    def determine_final_overdue_reason(row) -> str:
        if str(row.get("SLA Status", "")).strip() != "Overdue":
            return ""
        tid = str(row.get("Ticket ID", "")).strip()
        has_appt = tid in appt_dict
        has_spare = tid in spare_parts_tickets

        reason_part1 = ""
        if has_appt:
            reason_part1 = "Appointment"
        elif has_spare:
            part_info = ticket_to_part_str.get(tid, "")
            reason_part1 = f"Chờ vật tư {part_info}" if part_info else "Chờ vật tư"

        reason_part2 = ""
        if pd.notna(row.get("ASP close_dt")) and pd.notna(row.get("Create Time_dt")):
            time_to_first_asp = (row["ASP close_dt"] - row["Create Time_dt"]).total_seconds() / 3600
            if time_to_first_asp <= 48:
                if tid in reopened_tickets_set:
                    reason_part2 = "Ticket mở lại"
                else:
                    time_to_last_local = float("inf")
                    if pd.notna(row.get("Close Time_dt")):
                        time_to_last_local = (
                            row["Close Time_dt"] - row["Create Time_dt"]
                        ).total_seconds() / 3600
                    if time_to_last_local <= 48:
                        reason_part2 = "Lỗi chuyển trạng thái"
                    else:
                        reason_part2 = "ASP trong hạn, đóng Local Team trễ"
            else:
                try:
                    ticket_dur = float(row.get("Ticket Duration", 0) or 0)
                except (ValueError, TypeError):
                    ticket_dur = 0
                if ticket_dur > 48:
                    reason_part2 = "Tới site xử lý trong hạn nhưng case khó"

        if reason_part2:
            return f"{reason_part2} + {reason_part1}" if reason_part1 else reason_part2
        if reason_part1:
            return reason_part1
        return "Lỗi chủ quan"

    df_month_all["Nguyên nhân Overdue"] = df_month_all.apply(
        determine_final_overdue_reason, axis=1
    )

    target_columns = [
        "Kỹ thuật viên",
        "Ticket ID",
        "Create Time",
        "ASP close",
        "Close Time_dt",
        "Solution duration",
        "Station Code",
        "Charge Point ID",
        "Ticket Status",
        "Problem Description",
        "Address",
        "Handling type",
        "SLA Status",
        "Nguyên nhân Overdue",
        "Solution",
        "Giải trình",
        "Create Time_dt",
        "ASP close_dt",
        "First Local Close_dt",
        "Ticket Duration",
        "Khu vực",
    ]
    existing_target_cols = [c for c in target_columns if c in df_month_all.columns]

    report_date_folder_name = start_yesterday.strftime("%Y-%m-%d")
    print(f"\n☁️ Chuẩn bị thư mục Drive: {report_date_folder_name}")
    target_drive_folder_id = get_or_create_date_folder(DRIVE_FOLDER_ID, report_date_folder_name)
    print("-" * 60)

    df_unassigned = df_month_all[df_month_all["Kỹ thuật viên"] == "Chưa phân công"].copy()
    if not df_unassigned.empty:
        cols_unassigned = [
            c
            for c in existing_target_cols
            if c
            not in [
                "Kỹ thuật viên",
                "Solution",
                "Giải trình",
                "Create Time_dt",
                "Close Time_dt",
                "ASP close_dt",
                "First Local Close_dt",
                "Ticket Status",
                "Ticket Duration",
            ]
        ]
        xlsx_bytes = export_multisheet_excel(
            df_unassigned[cols_unassigned],
            None,
            None,
            None,
            yesterday_str,
            today_str,
            month,
            is_unassigned=True,
        )
        upload_bytes_to_drive(
            "unassigned_stations.xlsx", xlsx_bytes, target_drive_folder_id
        )
        print(f"   ✅ Unassigned: {len(df_unassigned)} ticket → Drive")

    df_all_parts_yesterday = get_replaced_parts_yesterday(
        df_month_all, df_spare, start_yesterday, end_today
    )

    for region in df_month_all["Region_op"].unique():
        if region in ("Chưa phân công", "UNKNOWN") or is_unmanaged_region(region):
            continue

        cols_region = list(existing_target_cols)
        if "Event_State" in df_month_all.columns and "Event_State" not in cols_region:
            cols_region.append("Event_State")
        df_region = df_month_all[df_month_all["Region_op"] == region][
            cols_region
        ].copy()
        if df_region.empty:
            continue

        # Cleaned = CLOSED theo Events; Report đếm từ cùng tập này.
        if "Event_State" in df_region.columns:
            df_region_cleaned = df_region[df_region["Event_State"] == "CLOSED"].copy()
        else:
            df_region_cleaned = df_region[
                ~df_region["Ticket Status"].isin(unclosed_statuses)
            ].copy()
        df_region_closed = df_region_cleaned

        report_df = generate_report_dataframe(
            df_region, df_region_closed, start_yesterday, end_today, unclosed_statuses
        )
        df_warning = check_missing_spare_parts(
            df_region_closed, df_events, df_spare, start_yesterday, end_today
        )
        df_region_parts = (
            df_all_parts_yesterday[
                df_all_parts_yesterday["Tên kỹ thuật"].isin(
                    set(report_df["Tên kỹ thuật"].unique())
                )
            ].copy()
            if not df_all_parts_yesterday.empty and report_df is not None and not report_df.empty
            else pd.DataFrame()
        )

        file_name = f"{region}.xlsx"

        if "Close Time_dt" in df_region_cleaned.columns:
            df_region_cleaned = df_region_cleaned.copy()
            df_region_cleaned["Close Time_dt"] = (
                pd.to_datetime(df_region_cleaned["Close Time_dt"], errors="coerce")
                .dt.strftime("%Y-%m-%d %H:%M:%S")
                .fillna("")
            )

        cols_to_save = [
            c
            for c in df_region_cleaned.columns
            if c
            not in [
                "Create Time_dt",
                "ASP close_dt",
                "First Local Close_dt",
                "Ticket Status",
                "Ticket Duration",
                "Event_State",
            ]
        ]
        if "Khu vực" in cols_to_save:
            cols_to_save = [c for c in cols_to_save if c != "Khu vực"] + ["Khu vực"]

        xlsx_bytes = export_multisheet_excel(
            df_region_cleaned[cols_to_save],
            report_df,
            df_warning,
            df_region_parts,
            yesterday_str,
            today_str,
            month,
            is_unassigned=False,
        )
        upload_bytes_to_drive(file_name, xlsx_bytes, target_drive_folder_id)
        print(f"   ✅ Vùng [{region}]: {len(df_region_cleaned)} ticket → Drive/{file_name}")

    print("\n🎉 HOÀN THÀNH BÁO CÁO NGÀY + ĐỒNG BỘ DRIVE!")
    return df_month_all


def run_daily_report_from_bundles(
    bundles: dict[str, dict[str, pd.DataFrame]],
    *,
    spare_parts_file: str = "spareParts.xlsx",
    now: datetime | None = None,
) -> pd.DataFrame | None:
    """
    bundles = {
        "esmanager": {sheet_name: DataFrame, ...},
        "itsmanagermt": {sheet_name: DataFrame, ...},
    }
    """
    dfs_es = bundles.get("esmanager") or bundles.get("es") or {}
    dfs_its = bundles.get("itsmanagermt") or bundles.get("its") or {}
    if not dfs_es and not dfs_its and bundles:
        items = list(bundles.values())
        dfs_es = items[0] if items else {}
        dfs_its = items[1] if len(items) > 1 else {}
    return process_and_clean_data(
        dfs_es, dfs_its, spare_parts_file=spare_parts_file, now=now
    )


if __name__ == "__main__":
    print(
        "[report] Chạy standalone — cần truyền dfs từ stats_data "
        "(hoặc import process_and_clean_data / run_daily_report_from_bundles)."
    )
    print("Ví dụ từ stats_data sau khi cào:")
    print("  from report import run_daily_report_from_bundles")
    print("  run_daily_report_from_bundles(bundles)")